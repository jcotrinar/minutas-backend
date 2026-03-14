"""
scripts/importar_excel.py — Importa datos del Excel a la nueva BD v2.
Uso:
  python scripts/importar_excel.py --excel MINUTAS.xlsm --proyecto 1
  python scripts/importar_excel.py --excel NSB.xlsm --proyecto 2
  python scripts/importar_excel.py --solo-distritos
"""
import sys, os, argparse
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from app.database import SessionLocal
from app.models import Lote, Distrito, Contrato, Moneda

# ─── MAPAS PARTIDA → NOMBRE PREDIO por proyecto ─────────────────────────────

# Proyecto 1 — Sol y Luna Malabrigo
PREDIO_SOL_LUNA = {
    "04020673": "PREDIO MOCAN LA ARENITA, VALLE CHICAMA, CON U.C. N° 1866",
    "11578607": "PREDIO MOCAN Y ANEXOS, PARCELA N° 1891",
    "04019384": "PREDIO RURAL VALLE CHICAMA FUNDO MOCAN, SECTOR LA ARENITA CON U.C. 1853",
    "04019387": "PREDIO RURAL VALLE CHICAMA FUNDO MOCAN, SECTOR LA ARENITA CON U.C. 1854",
    "04019390": "PREDIO MOCAN SECTOR LA ARENITA, PARCELA 1855, CON U.C. 1855",
    "04019385": "PREDIO FUNDO MOCAN-SECTOR LA ARENITA PARCELA 1856, CON U.C. 1856",
    "04017223": "PREDIO FUNDO MOCAN-SECTOR LA ARENITA, CON U.C. 1857",
    "04017224": "PREDIO FUNDO MOCAN-SECTOR LA ARENITA PARCELA, CON U.C. 1858",
    "11578604": "PREDIO MOCAN Y ANEXOS, PARCELA N° 1859",
    "04020058": "FUNDO MOCAN LA ARENITA, VALLE CHICAMA, CON U.C. 1865",
}

# Proyecto 2 — Santa Beatriz
PREDIO_SANTA_BEATRIZ = {
    "11579570": "UC 10399-B, PREDIO VALDIVIA, VALLE MOCHE",
    "04008773": "LOTE VD.80-III, SECTOR VALDIVIA BAJA, VALLE MOCHE",
}

# Proyecto 3 — Laureles
PREDIO_LAURELES = {
    "11095526": 'PREDIO RURAL DENOMINADO BLOCK "C", LOTE VD.144-III',
}

# Mapa global por proyecto_id
PREDIOS_POR_PROYECTO = {
    1: PREDIO_SOL_LUNA,
    2: PREDIO_SANTA_BEATRIZ,
    3: PREDIO_LAURELES,
}

def _es_error(val):
    return isinstance(val, str) and val.startswith("#")

def _sf(val, default=None):
    if val is None or _es_error(str(val) if val else ""): return default
    try: return float(val)
    except: return default

def _si(val, default=None):
    v = _sf(val)
    return int(v) if v is not None else default

def _ss(val, default=""):
    if val is None or _es_error(str(val)): return default
    return str(val).strip()

def _sd(val):
    if val is None: return None
    if hasattr(val, 'date'): return val.date()
    if hasattr(val, 'year'): return val
    return None


def importar_distritos(ws, db):
    print("Importando distritos...")
    if db.query(Distrito).count() > 0:
        print("  → Distritos ya importados, saltando.")
        return
    batch = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        region, provincia, distrito = row[0], row[1], row[2]
        if not region: continue
        batch.append(Distrito(
            region=_ss(region).upper(),
            provincia=_ss(provincia).upper(),
            distrito=_ss(distrito).upper(),
        ))
        if len(batch) >= 200:
            db.bulk_save_objects(batch); db.commit(); batch = []
    if batch:
        db.bulk_save_objects(batch); db.commit()
    print(f"  → {db.query(Distrito).count()} distritos importados")


def importar_lotes(ws, db, proyecto_id: int, predio_por_partida: dict = None):
    print(f"Importando lotes para proyecto {proyecto_id}...")
    db.query(Lote).filter(Lote.proyecto_id == proyecto_id).delete()
    db.commit()
    errores = 0
    lotes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        manzana = _ss(row[0])
        numero  = _ss(row[1])
        area    = _sf(row[2])
        partida = _ss(row[4]) if len(row) > 4 else ""
        if not manzana or not numero or area is None:
            errores += 1; continue
        nombre_predio = predio_por_partida.get(partida) if predio_por_partida and partida else None
        lotes.append(Lote(
            proyecto_id=proyecto_id, manzana=manzana, numero=numero,
            area=area, partida=partida or None, nombre_predio=nombre_predio,
        ))
        if len(lotes) >= 200:
            db.bulk_save_objects(lotes); db.commit(); lotes = []
    if lotes:
        db.bulk_save_objects(lotes); db.commit()
    print(f"  → {db.query(Lote).filter(Lote.proyecto_id == proyecto_id).count()} lotes importados ({errores} errores)")


def importar_contratos(ws, db, proyecto_id: int, moneda: str = "SOLES"):
    print(f"Importando contratos para proyecto {proyecto_id}...")
    db.query(Contrato).filter(Contrato.proyecto_id == proyecto_id).delete()
    db.commit()

    importados = 0
    errores    = 0

    # Leer todas las filas primero para invertir el orden
    # Así el contrato más reciente del Excel (última fila) queda con numero_proyecto más alto
    todas_filas = [row for row in ws.iter_rows(min_row=6, values_only=True) if row[0]]
    todas_filas.reverse()  # invertir: última fila del Excel = numero_proyecto 1, primera = el mayor

    numero_proyecto = 1

    for row in todas_filas:
        if not row[0]: continue

        fecha      = _sd(row[1])
        mz         = _ss(row[2])
        lote_num   = _ss(row[3])
        precio     = _sf(row[6], 0.0)
        titular    = _ss(row[9])
        ocupacion1 = _ss(row[10])
        genero1    = _ss(row[11])
        estado_c1  = _ss(row[12])
        dni        = _ss(row[13])
        direccion1 = _ss(row[14])
        copropiet  = _ss(row[15])
        ocupacion2 = _ss(row[16])
        genero2    = _ss(row[17])
        estado_c2  = _ss(row[18])
        dni2       = _ss(row[19])
        direccion2 = _ss(row[20])
        separacion = _sf(row[21], 0.0)
        pago       = _sf(row[22], 0.0)
        f_sep      = _sd(row[24]) if len(row) > 24 else None

        if not titular or not fecha: continue

        lote = db.query(Lote).filter(
            Lote.proyecto_id == proyecto_id,
            Lote.manzana == mz,
            Lote.numero == lote_num
        ).first()

        if not lote:
            errores += 1; continue

        c = Contrato(
            numero_proyecto=numero_proyecto,   # ← correlativo secuencial por proyecto
            proyecto_id=proyecto_id,
            lote_id=lote.id,
            fecha=fecha,
            titular=titular.upper(),
            dni=dni or None,
            ocupacion1=ocupacion1 or None,
            genero1=genero1 or None,
            estado_civil1=estado_c1 or None,
            direccion1=direccion1 or None,
            copropietario=copropiet.upper() if copropiet else None,
            ocupacion2=ocupacion2 or None,
            genero2=genero2 or None,
            estado_civil2=estado_c2 or None,
            dni2=dni2 or None,
            direccion2=direccion2 or None,
            moneda=Moneda(moneda),
            precio=precio,
            separacion=separacion,
            pago=pago,
            f_pago_total=f_sep,
        )
        db.add(c)
        numero_proyecto += 1
        importados += 1

        if importados % 100 == 0:
            db.commit()
            print(f"  ... {importados} contratos")

    db.commit()
    print(f"  → {importados} contratos importados ({errores} errores)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel",          help="Ruta al archivo .xlsm")
    parser.add_argument("--proyecto",       type=int, help="ID del proyecto en BD")
    parser.add_argument("--moneda",         default="SOLES", choices=["SOLES", "DOLARES"])
    parser.add_argument("--solo-distritos", action="store_true")
    args = parser.parse_args()

    db = SessionLocal()
    try:
        if args.solo_distritos:
            wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
            importar_distritos(wb["DISTRITOS"], db)
            return

        if not args.excel or not args.proyecto:
            print("ERROR: Se requiere --excel y --proyecto")
            sys.exit(1)

        print(f"Abriendo {args.excel}...")
        wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)

        if "DISTRITOS" in wb.sheetnames:
            importar_distritos(wb["DISTRITOS"], db)

        if "LOTES" in wb.sheetnames:
            # Usar el mapa de predios correspondiente al proyecto
            predio_map = PREDIOS_POR_PROYECTO.get(args.proyecto)
            importar_lotes(wb["LOTES"], db, args.proyecto, predio_map)

        hoja_contratos = "Hoja1" if "Hoja1" in wb.sheetnames else wb.sheetnames[0]
        importar_contratos(wb[hoja_contratos], db, args.proyecto, args.moneda)

        print("\n✓ Importación completada.")
    finally:
        db.close()

if __name__ == "__main__":
    main()